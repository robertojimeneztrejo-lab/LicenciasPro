import streamlit as st
import google.generativeai as genai
import json
import re
import io
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Buscador Licencias",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── CUSTOM CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600&display=swap');

:root {
    --teal:    #0d9488;
    --teal-lt: #14b8a6;
    --navy:    #0f172a;
    --slate:   #1e293b;
    --muted:   #475569;
    --border:  #1e3a5f;
    --card:    #0f2137;
    --amber:   #f59e0b;
}

html, body, [data-testid="stAppViewContainer"] {
    background: var(--navy) !important;
    font-family: 'DM Sans', sans-serif;
    color: #e2e8f0;
}
#MainMenu, footer, header { visibility: hidden; }
[data-testid="stToolbar"] { display: none; }

/* Hero */
.hero {
    background: #0c3b4a;
    border: 1px solid var(--border);
    border-radius: 16px;
    padding: 2.2rem 2.8rem;
    margin-bottom: 2rem;
}
.hero-badge {
    display: inline-block;
    background: rgba(13,148,136,.2);
    border: 1px solid var(--teal);
    color: var(--teal-lt);
    font-size: .72rem; font-weight: 600;
    letter-spacing: .12em; text-transform: uppercase;
    padding: .3rem .8rem; border-radius: 99px; margin-bottom: 1rem;
}
.hero h1 {
    font-family: 'DM Serif Display', serif;
    font-size: 2.4rem; line-height: 1.15;
    color: #f1f5f9; margin: 0 0 .7rem; font-weight: 400;
}
.hero h1 span { color: var(--teal-lt); }
.hero p { color: #94a3b8; font-size: .95rem; max-width: 620px; line-height: 1.65; margin: 0; }

/* Stage pills */
.stage-row { display: flex; gap: .75rem; margin-bottom: 1.5rem; flex-wrap: wrap; }
.stage-pill {
    display: flex; align-items: center; gap: .5rem;
    background: var(--card); border: 1px solid var(--border);
    border-radius: 99px; padding: .4rem .9rem;
    font-size: .8rem; font-weight: 500; color: #94a3b8;
}
.stage-pill.active { background: rgba(13,148,136,.15); border-color: var(--teal); color: var(--teal-lt); }
.stage-dot { width: 7px; height: 7px; border-radius: 50%; background: currentColor; opacity: .6; }

/* Config label */
.config-label {
    font-size: .75rem; font-weight: 600; letter-spacing: .1em;
    text-transform: uppercase; color: var(--muted); margin-bottom: .45rem;
}

/* Inputs */
[data-testid="stTextInput"] input,
[data-testid="stTextArea"] textarea,
[data-testid="stSelectbox"] select {
    background: #0a1929 !important;
    border: 1px solid var(--border) !important;
    border-radius: 8px !important;
    color: #e2e8f0 !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stTextArea"] textarea:focus {
    border-color: var(--teal) !important;
    box-shadow: 0 0 0 2px rgba(13,148,136,.25) !important;
}

/* Buttons */
.stButton > button {
    background: linear-gradient(135deg, var(--teal) 0%, #0d7a70 100%) !important;
    color: white !important; font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important; font-size: .88rem !important;
    border: none !important; border-radius: 10px !important;
    padding: .6rem 1.6rem !important; transition: all .2s !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(13,148,136,.35) !important;
}

/* Correction button override — amber */
.stButton[data-testid*="btn_apply"] > button,
button[kind="secondary"] {
    background: transparent !important;
    border: 1px solid rgba(245,158,11,.5) !important;
    color: #f59e0b !important;
}

/* Download button */
.stDownloadButton > button {
    background: linear-gradient(135deg, #065f46 0%, #047857 100%) !important;
    color: white !important; font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important; border: 1px solid #059669 !important;
    border-radius: 10px !important; padding: .6rem 1.6rem !important;
}
.stDownloadButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(5,150,105,.35) !important;
}

/* Status boxes */
.status-box {
    background: var(--card); border: 1px solid var(--border);
    border-left: 3px solid var(--teal); border-radius: 0 8px 8px 0;
    padding: .9rem 1.1rem; font-size: .88rem; color: #94a3b8; margin: .8rem 0;
}

/* Feedback box (amber) */
.feedback-box {
    background: rgba(245,158,11,.05);
    border: 1px solid rgba(245,158,11,.3);
    border-left: 3px solid var(--amber);
    border-radius: 0 10px 10px 0;
    padding: 1.2rem 1.4rem; margin: 1.2rem 0;
}
.feedback-title {
    font-size: .75rem; font-weight: 700; letter-spacing: .1em;
    text-transform: uppercase; color: var(--amber); margin-bottom: .5rem;
}
.feedback-body { font-size: .88rem; color: #cbd5e1; line-height: 1.55; margin: 0; }

/* Correction applied box */
.correction-box {
    background: rgba(13,148,136,.06);
    border: 1px solid rgba(13,148,136,.3);
    border-left: 3px solid var(--teal-lt);
    border-radius: 0 8px 8px 0;
    padding: .9rem 1.1rem; margin: .8rem 0;
    font-size: .88rem; color: #94a3b8;
}

/* Result cards */
.result-card {
    background: var(--card); border: 1px solid var(--border);
    border-radius: 14px; padding: 1.4rem; margin-bottom: .9rem;
    transition: border-color .2s;
}
.result-card:hover { border-color: var(--teal); }
.result-num {
    display: inline-flex; align-items: center; justify-content: center;
    width: 26px; height: 26px;
    background: rgba(13,148,136,.2); border: 1px solid var(--teal);
    border-radius: 50%; font-size: .78rem; font-weight: 700;
    color: var(--teal-lt); margin-bottom: .65rem;
}
.result-name { font-family: 'DM Serif Display', serif; font-size: 1.15rem; color: #f1f5f9; margin-bottom: .25rem; }
.result-url { font-size: .8rem; color: var(--teal-lt); margin-bottom: .9rem; word-break: break-all; }
.result-field { margin-bottom: .7rem; }
.result-field-label { font-size: .7rem; font-weight: 700; letter-spacing: .1em; text-transform: uppercase; color: var(--muted); margin-bottom: .22rem; }
.result-field-value { font-size: .86rem; color: #cbd5e1; line-height: 1.55; }

/* Block header */
.block-header {
    display: flex; align-items: center; gap: .6rem; margin-bottom: .8rem;
}
.block-badge {
    background: rgba(13,148,136,.15); border: 1px solid rgba(13,148,136,.3);
    color: var(--teal-lt); font-size: .7rem; font-weight: 600;
    padding: .22rem .7rem; border-radius: 99px;
}
.block-range { font-size: .8rem; color: var(--muted); }

/* More block */
.more-block {
    background: rgba(13,148,136,.04);
    border: 1px dashed rgba(13,148,136,.35);
    border-radius: 12px; padding: 1.5rem 1.8rem;
    margin: 1.2rem 0; text-align: center;
}
.more-badge {
    display: inline-block;
    background: rgba(13,148,136,.15); border: 1px solid rgba(13,148,136,.3);
    color: var(--teal-lt); font-size: .68rem; font-weight: 600;
    letter-spacing: .1em; text-transform: uppercase;
    padding: .22rem .7rem; border-radius: 99px; margin-bottom: .6rem;
}
.more-title { font-size: .95rem; color: #e2e8f0; font-weight: 500; margin-bottom: .3rem; }
.more-sub { font-size: .82rem; color: #64748b; margin-bottom: .9rem; }

/* Divider */
.divider { height: 1px; background: linear-gradient(90deg, transparent, var(--border), transparent); margin: 1.5rem 0; }

/* Metrics */
[data-testid="stMetric"] {
    background: var(--card) !important; border: 1px solid var(--border) !important;
    border-radius: 10px !important; padding: 1rem !important;
}
[data-testid="stMetricValue"] { color: var(--teal-lt) !important; font-family: 'DM Serif Display', serif !important; }
[data-testid="stMetricLabel"] { color: var(--muted) !important; font-size: .75rem !important; }

/* Expander */
[data-testid="stExpander"] {
    background: var(--card) !important; border: 1px solid var(--border) !important;
    border-radius: 10px !important;
}

/* Sidebar */
[data-testid="stSidebar"] { background: var(--slate) !important; border-right: 1px solid var(--border) !important; }

::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: var(--navy); }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--teal); }
</style>
""", unsafe_allow_html=True)


# ─── PROMPTS ──────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Eres un analista de software experto especializado en herramientas para el sector académico y universitario.
Tu tarea es encontrar herramientas digitales (softwares, web apps o apps) siguiendo un proceso de 3 etapas estrictas.

REGLAS DE RESPUESTA:
- Responde SIEMPRE en JSON válido y sin texto extra.
- No incluyas bloques de código (```json), solo el JSON puro.
- Sé preciso, verificable y concreto en cada campo.
"""

ETAPA1_PROMPT = """## ETAPA 1: FILTRO DE MODELO DE NEGOCIO Y ACCESO (Inclusión)

Tema de búsqueda: {tema}

Busca herramientas que cumplan OBLIGATORIAMENTE:
1. Es de PAGO para el público general (comercial/premium).
2. Es 100% GRATUITO para el sector educativo (universidades, académicos, instituciones o estudiantes) — sin costo para el estudiante NI para su institución.
3. El acceso gratuito educativo NO es automático: exige validación humana o institucional (formulario, cuestionario, correo .edu, aprobación manual).
4. EXCLUIR si solo ofrecen descuento (aunque sea 90% off).
5. EXCLUIR si la gratuidad depende de que la universidad haya comprado una licencia institucional.

REQUISITO CLAVE: El software debe exigir que el académico llene un formulario, cuestionario o envíe correo institucional .edu para que LE APRUEBEN manualmente la licencia gratuita. Busca páginas con textos como: "Apply for Academic Grant", "Request Educational License", "Academic Application Form", "Contact Support for Education".

NOTA VÁLIDA: Si el software tiene plan básico gratis para todos, pero regala el plan PRO/PAGO a estudiantes/educadores mediante este proceso de verificación manual, SÍ es válido.

Proporciona exactamente 5 ejemplos de validación que cumplan este modelo para el tema "{tema}".

Responde en este JSON exacto:
{{
  "etapa": 1,
  "tema": "{tema}",
  "herramientas_validacion": [
    {{
      "nombre": "...",
      "url": "...",
      "descripcion_breve": "...",
      "evidencia_pago": "...",
      "evidencia_acceso_educativo": "..."
    }}
  ]
}}"""

ETAPA1_CORRECCION_PROMPT = """## CORRECCIÓN DE ETAPA 1

Resultados anteriores de Etapa 1:
{resultados_anteriores}

El usuario ha revisado los resultados y tiene estas correcciones o instrucciones adicionales:
"{instrucciones_usuario}"

Aplica las correcciones indicadas:
- Si el usuario dice que alguna herramienta NO cumple el filtro, elimínala y reemplázala por una que sí cumpla.
- Si el usuario pide agregar herramientas específicas, agrégalas verificando que cumplan el filtro.
- Si el usuario pide cambiar criterios, aplica el cambio y regenera la lista.
- Mantén exactamente 5 herramientas en la lista final corregida.

Responde en este JSON exacto (mismo formato que Etapa 1):
{{
  "etapa": 1,
  "tema": "{tema}",
  "herramientas_validacion": [
    {{
      "nombre": "...",
      "url": "...",
      "descripcion_breve": "...",
      "evidencia_pago": "...",
      "evidencia_acceso_educativo": "..."
    }}
  ]
}}"""

ETAPA2_PROMPT = """## ETAPA 2: FILTRO DE VISIBILIDAD Y EXCLUSIÓN (Emergentes)

Herramientas pre-seleccionadas en Etapa 1: {herramientas_etapa1}

Aplica este filtro RADICAL sobre las herramientas anteriores y busca herramientas SIMILARES pero emergentes:

1. DESCARTA software Open Source o proyectos 100% gratuitos para el público general.
2. DESCARTA softwares populares, masivos o de empresas líderes (NO Adobe, NO Autodesk, NO JetBrains, NO Notion, NO Canva, NO Microsoft, NO grandes corporativos).
3. REQUISITO DE BAJA POPULARIDAD: Solo herramientas recientes, startups emergentes, softwares en fase beta pública o páginas con muy baja interacción en el mercado. Deben ser herramientas que NO aparezcan en los primeros 3-5 resultados de Google.
4. Busca en las capas más profundas y menos comerciales de la web.

Genera exactamente {cantidad} herramientas emergentes para "{tema}" que pasen AMBOS filtros.
Empieza la numeración desde {num_inicio}.

Responde en este JSON exacto:
{{
  "etapa": 2,
  "tema": "{tema}",
  "herramientas_filtradas": [
    {{
      "numero": {num_inicio},
      "nombre": "...",
      "url": "...",
      "categoria_funcion": "...",
      "precio_publico": "...",
      "metodo_acceso_educativo": "...",
      "razon_baja_popularidad": "..."
    }}
  ]
}}"""

ETAPA3_PROMPT = """## ETAPA 3: LISTA DEFINITIVA ENRIQUECIDA

Herramientas encontradas: {herramientas_etapa2}

Para cada herramienta, enriquece la información con:
1. Descripción detallada de funcionalidades clave.
2. Evidencia concreta de que es de pago (precio exacto si existe, o modelo de facturación).
3. Proceso EXACTO de validación educativa (pasos, formulario URL si existe, qué datos piden).
4. Por qué califica como emergente/poco conocida.
5. Casos de uso específicos para estudiantes/docentes del área.

Responde en este JSON exacto:
{{
  "etapa": 3,
  "tema": "{tema}",
  "lista_definitiva": [
    {{
      "numero": 0,
      "nombre": "...",
      "url": "...",
      "que_hace_categoria": "...",
      "precio_publico_general": "...",
      "metodo_acceso_educativo_gratuito": "...",
      "razon_baja_popularidad": "...",
      "casos_uso": "..."
    }}
  ]
}}"""


# ─── GEMINI HELPERS ───────────────────────────────────────────────────────────

def get_api_key() -> str:
    try:
        return st.secrets["GEMINI_API_KEY"]
    except (KeyError, FileNotFoundError):
        return ""


def get_model(api_key: str):
    genai.configure(api_key=api_key)
    return genai.GenerativeModel(
        model_name="gemini-2.5-flash",
        system_instruction=SYSTEM_PROMPT,
    )


def call_gemini(model, prompt: str, retries: int = 3) -> dict:
    for attempt in range(retries):
        try:
            response = model.generate_content(prompt)
            text = response.text.strip()
            text = re.sub(r"^```(?:json)?\s*", "", text)
            text = re.sub(r"\s*```$", "", text)
            return json.loads(text)
        except json.JSONDecodeError:
            if attempt == retries - 1:
                return {"error": "JSON inválido en respuesta de Gemini", "raw": text}
        except Exception as e:
            if attempt == retries - 1:
                return {"error": str(e)}
            time.sleep(2 ** attempt)
    return {"error": "Sin respuesta"}


# ─── EXCEL BUILDER ────────────────────────────────────────────────────────────

HEADERS = [
    "N°", "Nombre", "URL", "Categoría / Función",
    "Precio Público General", "Método Acceso Educativo Gratuito",
    "Razón Baja Popularidad", "Casos de Uso",
]
COL_WIDTHS = [5, 28, 35, 30, 30, 45, 40, 45]
TEAL  = "0D9488"
NAVY  = "0F172A"
SLATE = "1E293B"
WHITE = "F1F5F9"
LIGHT = "E2E8F0"


def _border():
    s = Side(style="thin", color="1E3A5F")
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel(results: list, tema: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"Buscador Licencias — {tema}"
    c.font = Font(name="Arial", bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = "Softwares académicos con alto impacto · Pago público | Gratuito académico verificado | Emergentes"
    s.font = Font(name="Arial", size=9, color="94A3B8")
    s.fill = PatternFill("solid", start_color=NAVY)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Spacer
    for col in range(1, 9):
        ws.cell(row=3, column=col).fill = PatternFill("solid", start_color=NAVY)
    ws.row_dimensions[3].height = 6

    # Header row
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        cell.fill = PatternFill("solid", start_color=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 28

    # Data rows
    for i, item in enumerate(results):
        row = 5 + i
        bg = NAVY if i % 2 == 0 else SLATE
        values = [
            item.get("numero", i + 1),
            item.get("nombre", ""),
            item.get("url", ""),
            item.get("que_hace_categoria", item.get("categoria_funcion", "")),
            item.get("precio_publico_general", item.get("precio_publico", "")),
            item.get("metodo_acceso_educativo_gratuito", item.get("metodo_acceso_educativo", "")),
            item.get("razon_baja_popularidad", ""),
            item.get("casos_uso", item.get("casos_uso_enfermeria", "")),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=str(value))
            cell.fill = PatternFill("solid", start_color=bg)
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal="center" if col_idx == 1 else "left")
            if col_idx == 1:
                cell.font = Font(name="Arial", bold=True, size=10, color="0D9488")
            elif col_idx == 2:
                cell.font = Font(name="Arial", bold=True, size=9, color=WHITE)
            elif col_idx == 3:
                cell.font = Font(name="Arial", size=8, color="14B8A6", underline="single")
            else:
                cell.font = Font(name="Arial", size=9, color=LIGHT)
        ws.row_dimensions[row].height = 80

    ws.freeze_panes = "A5"

    # Summary sheet
    ws2 = wb.create_sheet("Resumen")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 55

    total_bloques = max(1, (len(results) + 9) // 10)
    summary_data = [
        ("Tema de búsqueda", tema),
        ("Total herramientas", len(results)),
        ("Bloques generados", total_bloques),
        ("Etapas aplicadas", "3 (Modelo de negocio + Visibilidad + Enriquecimiento)"),
        ("Filtro 1", "De pago para público general"),
        ("Filtro 2", "100% gratuito academia con validación manual"),
        ("Filtro 3", "Emergente / baja popularidad / no en top 5 Google"),
        ("Motor IA", "Gemini 2.5 Flash"),
        ("Generado con", "Buscador Licencias v1.2"),
    ]

    ws2.merge_cells("A1:B1")
    t = ws2["A1"]
    t.value = "Resumen de Búsqueda"
    t.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    t.fill = PatternFill("solid", start_color=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    for row_idx, (label, value) in enumerate(summary_data, start=2):
        bg = NAVY if row_idx % 2 == 0 else SLATE
        lc = ws2.cell(row=row_idx, column=1, value=label)
        lc.font = Font(name="Arial", bold=True, size=9, color="94A3B8")
        lc.fill = PatternFill("solid", start_color=bg)
        lc.border = _border()
        lc.alignment = Alignment(vertical="center")
        vc = ws2.cell(row=row_idx, column=2, value=str(value))
        vc.font = Font(name="Arial", size=9, color=WHITE)
        vc.fill = PatternFill("solid", start_color=bg)
        vc.border = _border()
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        ws2.row_dimensions[row_idx].height = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ─── UI HELPERS ───────────────────────────────────────────────────────────────

def render_hero():
    st.markdown("""
    <div class="hero">
        <div class="hero-badge">🎓 Licencias Académicas Gratuitas</div>
        <h1>Buscador de <span>Licencias</span></h1>
        <p>Buscador de softwares académicos con alto impacto en el desarrollo académico
        de la comunidad Universitaria.</p>
    </div>
    """, unsafe_allow_html=True)


def render_stage_indicator(active: int):
    stages = [("1","Modelo de Negocio"), ("2","Filtro Emergentes"), ("3","Lista Definitiva")]
    pills = ""
    for num, label in stages:
        cls = "stage-pill active" if int(num) <= active else "stage-pill"
        pills += f'<div class="{cls}"><div class="stage-dot"></div> Etapa {num}: {label}</div>'
    st.markdown(f'<div class="stage-row">{pills}</div>', unsafe_allow_html=True)


def render_result_card(item: dict):
    num      = item.get("numero", "?")
    nombre   = item.get("nombre", "Sin nombre")
    url      = item.get("url", "#")
    categoria= item.get("que_hace_categoria", item.get("categoria_funcion", ""))
    precio   = item.get("precio_publico_general", item.get("precio_publico", ""))
    acceso   = item.get("metodo_acceso_educativo_gratuito", item.get("metodo_acceso_educativo", ""))
    popular  = item.get("razon_baja_popularidad", "")
    casos    = item.get("casos_uso", item.get("casos_uso_enfermeria", ""))
    casos_html = (f'<div class="result-field"><div class="result-field-label">🏥 Casos de Uso</div>'
                  f'<div class="result-field-value">{casos}</div></div>') if casos else ""
    st.markdown(f"""
    <div class="result-card">
        <div class="result-num">{num}</div>
        <div class="result-name">{nombre}</div>
        <div class="result-url">🔗 {url}</div>
        <div class="result-field">
            <div class="result-field-label">📋 Categoría / Función</div>
            <div class="result-field-value">{categoria}</div>
        </div>
        <div class="result-field">
            <div class="result-field-label">💰 Precio Público General</div>
            <div class="result-field-value">{precio}</div>
        </div>
        <div class="result-field">
            <div class="result-field-label">🎓 Acceso Educativo Gratuito</div>
            <div class="result-field-value">{acceso}</div>
        </div>
        <div class="result-field">
            <div class="result-field-label">🔍 Por qué es Emergente</div>
            <div class="result-field-value">{popular}</div>
        </div>
        {casos_html}
    </div>
    """, unsafe_allow_html=True)


def render_block_header(bloque_num: int, start: int, end: int):
    st.markdown(f"""
    <div class="block-header">
        <div class="block-badge">Bloque {bloque_num}</div>
        <span class="block-range">#{start} — #{end}</span>
    </div>
    """, unsafe_allow_html=True)


def render_e1_expander(herramientas: list):
    with st.expander(f"✅ Etapa 1 — {len(herramientas)} herramientas pre-seleccionadas (revisa antes de continuar)"):
        for h in herramientas:
            st.markdown(f"**{h.get('nombre')}** — `{h.get('url')}`")
            col_a, col_b = st.columns(2)
            with col_a:
                st.caption(f"💰 {h.get('evidencia_pago','')}")
            with col_b:
                st.caption(f"🎓 {h.get('evidencia_acceso_educativo','')}")
            st.markdown("---")


# ─── MAIN APP ─────────────────────────────────────────────────────────────────

def main():

    # ── Session state init — MUST be first, before any widget or logic ───
    for key, default in [
        ("e1_done", False),
        ("e1_result", None),
        ("model", None),
        ("corrected", False),
        ("correction_applied", False),
        ("all_results", []),
        ("bloque_num", 0),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    render_hero()

    api_key = get_api_key()

    # Sidebar
    with st.sidebar:
        st.markdown("### 🎓 Buscador Licencias")
        st.markdown("""
        **Flujo de búsqueda:**

        1. **Etapa 1** — Valida modelo de negocio
        2. ✏️ **Corrección opcional**
        3. **Etapa 2** — Filtra emergentes
        4. **Etapa 3** — Lista definitiva
        5. **＋** Busca más bloques
        6. **Descarga** Excel acumulado
        """)
        st.markdown("---")
        st.markdown("<small style='color:#475569'>Buscador Licencias v1.2<br>Powered by Gemini 2.5 Flash</small>",
                    unsafe_allow_html=True)

    # Config inputs
    col_left, col_right = st.columns([2, 1])
    with col_left:
        st.markdown('<div class="config-label">Tema de búsqueda</div>', unsafe_allow_html=True)
        tema = st.text_input(
            label="tema", label_visibility="collapsed",
            value="",
            placeholder="Ej: Enfermería, Arquitectura, Marketing Digital, Derecho...",
        )
    with col_right:
        st.markdown('<div class="config-label">Resultados por bloque</div>', unsafe_allow_html=True)
        cantidad = st.selectbox(
            label="cantidad", label_visibility="collapsed",
            options=[5, 10, 15, 20], index=1,
        )

    run = st.button("🔍 Buscar Herramientas", use_container_width=True)

    # Validate
    if (run or st.session_state.e1_done) and not api_key:
        st.error("⚠️ No se encontró la API Key. Agrega `GEMINI_API_KEY` en **Settings → Secrets** de Streamlit Cloud.")
        return

    # ── FRESH RUN ─────────────────────────────────────────────────────────
    if run:
        if not tema.strip():
            st.error("⚠️ Ingresa un tema de búsqueda.")
            return
        # Reset everything
        for key, val in [
            ("e1_done", False), ("e1_result", None), ("model", None),
            ("corrected", False), ("correction_applied", False),
            ("all_results", []), ("bloque_num", 0),
        ]:
            st.session_state[key] = val

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        try:
            model = get_model(api_key)
            st.session_state.model = model
        except Exception as e:
            st.error(f"Error al configurar Gemini: {e}")
            return

        render_stage_indicator(1)
        with st.spinner("Ejecutando Etapa 1..."):
            st.markdown('<div class="status-box">⚙️ Etapa 1 — Analizando modelos de negocio y acceso educativo...</div>',
                        unsafe_allow_html=True)
            result1 = call_gemini(model, ETAPA1_PROMPT.format(tema=tema))

        if "error" in result1:
            st.error(f"Error en Etapa 1: {result1['error']}")
            return

        st.session_state.e1_result = result1
        st.session_state.e1_done = True
        st.rerun()

    # ── ETAPA 1 DONE: show results + feedback ────────────────────────────
    if st.session_state.e1_done and st.session_state.e1_result:

        result1       = st.session_state.e1_result
        herramientas_e1 = result1.get("herramientas_validacion", [])
        model         = st.session_state.model or get_model(api_key)
        st.session_state.model = model
        # Recover tema from result
        tema = result1.get("tema", tema)

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        render_stage_indicator(1)
        render_e1_expander(herramientas_e1)

        # ── FEEDBACK FORM (shown until correction_applied) ────────────────
        if not st.session_state.correction_applied:
            st.markdown("""
            <div class="feedback-box">
                <div class="feedback-title">✏️ Revisión opcional antes de continuar</div>
                <div class="feedback-body">
                    Revisa los resultados de Etapa 1. Si alguna herramienta no cumple el filtro
                    o quieres instrucciones adicionales, escríbelas aquí.
                    Si todo está correcto, déjalo vacío y haz clic en <strong>Continuar a Etapa 2</strong>.
                </div>
            </div>
            """, unsafe_allow_html=True)

            instrucciones = st.text_area(
                label="Instrucciones de corrección",
                label_visibility="collapsed",
                placeholder=(
                    'Ej: "Elimina NVivo porque solo ofrece descuento, no acceso gratuito completo. '
                    'Reemplázalo por una herramienta de simulación clínica emergente." '
                    '— O déjalo vacío para continuar sin cambios.'
                ),
                height=110,
                key="feedback_text",
            )

            col_corr, col_cont = st.columns(2)
            with col_corr:
                apply_correction = st.button(
                    "✏️ Aplicar Corrección",
                    use_container_width=True,
                    disabled=not instrucciones.strip(),
                    key="btn_apply_correction",
                )
            with col_cont:
                continue_btn = st.button(
                    "▶ Continuar a Etapa 2",
                    use_container_width=True,
                    key="btn_continue",
                )

            if apply_correction and instrucciones.strip():
                with st.spinner("Aplicando correcciones..."):
                    result_corr = call_gemini(model, ETAPA1_CORRECCION_PROMPT.format(
                        resultados_anteriores=json.dumps(herramientas_e1, ensure_ascii=False),
                        instrucciones_usuario=instrucciones.strip(),
                        tema=tema,
                    ))
                if "error" in result_corr:
                    st.error(f"Error al aplicar corrección: {result_corr['error']}")
                else:
                    st.session_state.e1_result = result_corr
                    st.session_state.corrected = True
                    st.markdown('<div class="correction-box">✅ Corrección aplicada. Revisa los resultados y haz clic en <strong>Continuar a Etapa 2</strong>.</div>',
                                unsafe_allow_html=True)
                    st.rerun()

            if not continue_btn:
                return

            st.session_state.correction_applied = True

        # ── ETAPAS 2 + 3 (primer bloque) ─────────────────────────────────
        herramientas_e1 = st.session_state.e1_result.get("herramientas_validacion", [])
        nombres_e1      = [h.get("nombre", "") for h in herramientas_e1]
        num_inicio      = len(st.session_state.all_results) + 1

        render_stage_indicator(2)
        with st.spinner("Etapa 2 — Filtrando emergentes..."):
            st.markdown('<div class="status-box">🔍 Etapa 2 — Filtrando por emergencia y baja popularidad...</div>',
                        unsafe_allow_html=True)
            result2 = call_gemini(model, ETAPA2_PROMPT.format(
                herramientas_etapa1=json.dumps(nombres_e1, ensure_ascii=False),
                cantidad=cantidad,
                tema=tema,
                num_inicio=num_inicio,
            ))

        if "error" in result2:
            st.error(f"Error en Etapa 2: {result2['error']}")
            st.session_state.correction_applied = False
            return

        herramientas_e2 = result2.get("herramientas_filtradas", [])
        with st.expander(f"✅ Etapa 2 — {len(herramientas_e2)} herramientas emergentes seleccionadas"):
            for h in herramientas_e2:
                st.markdown(f"- **{h.get('nombre')}** — `{h.get('url')}`")
                st.caption(h.get("razon_baja_popularidad", ""))

        render_stage_indicator(3)
        with st.spinner("Etapa 3 — Enriqueciendo información..."):
            st.markdown('<div class="status-box">📋 Etapa 3 — Enriqueciendo información detallada...</div>',
                        unsafe_allow_html=True)
            result3 = call_gemini(model, ETAPA3_PROMPT.format(
                herramientas_etapa2=json.dumps(herramientas_e2, ensure_ascii=False),
                tema=tema,
            ))

        if "error" in result3:
            st.error(f"Error en Etapa 3: {result3['error']}")
            st.session_state.correction_applied = False
            return

        lista_bloque = result3.get("lista_definitiva", [])
        for i, item in enumerate(lista_bloque):
            item["numero"] = num_inicio + i

        st.session_state.all_results.extend(lista_bloque)
        st.session_state.bloque_num += 1
        st.session_state.correction_applied = False
        st.rerun()

    # ── SHOW ACCUMULATED RESULTS ──────────────────────────────────────────
    if st.session_state.all_results:
        all_results = st.session_state.all_results
        tema        = st.session_state.e1_result.get("tema", "")
        model       = st.session_state.model or get_model(api_key)
        st.session_state.model = model

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # Metrics
        m1, m2, m3 = st.columns(3)
        m1.metric("Total encontradas", len(all_results))
        m2.metric("Bloques generados", st.session_state.bloque_num)
        end_num = all_results[-1].get("numero", len(all_results)) if all_results else 0
        m3.metric("Rango actual", f"#1 → #{end_num}")

        if st.session_state.corrected:
            st.markdown('<div class="correction-box">🔧 Lista generada con correcciones aplicadas en Etapa 1.</div>',
                        unsafe_allow_html=True)

        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # Render blocks
        block_size = cantidad
        for b in range(st.session_state.bloque_num):
            bloque_items = all_results[b * block_size: (b + 1) * block_size]
            if not bloque_items:
                continue
            start_n = bloque_items[0].get("numero", b * block_size + 1)
            end_n   = bloque_items[-1].get("numero", (b + 1) * block_size)
            render_block_header(b + 1, start_n, end_n)
            for item in bloque_items:
                render_result_card(item)
            if b < st.session_state.bloque_num - 1:
                st.markdown('<div class="divider"></div>', unsafe_allow_html=True)

        # ── "Buscar más" button ───────────────────────────────────────────
        next_bloque = st.session_state.bloque_num + 1
        next_start  = end_num + 1
        next_end    = next_start + cantidad - 1

        st.markdown(f"""
        <div class="more-block">
            <div class="more-badge">Continuar búsqueda</div>
            <div class="more-title">¿Quieres más herramientas?</div>
            <div class="more-sub">Se generará el Bloque {next_bloque} (#{next_start}–#{next_end}) sin perder los resultados anteriores</div>
        </div>
        """, unsafe_allow_html=True)

        buscar_mas = st.button(
            f"＋ Buscar {cantidad} más (Bloque {next_bloque})",
            use_container_width=False,
            key=f"btn_more_{st.session_state.bloque_num}",
        )

        if buscar_mas:
            nombres_e1 = [h.get("nombre","") for h in
                          st.session_state.e1_result.get("herramientas_validacion", [])]
            num_inicio = next_start

            with st.spinner(f"Generando Bloque {next_bloque}..."):
                result2 = call_gemini(model, ETAPA2_PROMPT.format(
                    herramientas_etapa1=json.dumps(nombres_e1, ensure_ascii=False),
                    cantidad=cantidad,
                    tema=tema,
                    num_inicio=num_inicio,
                ))

            if "error" in result2:
                st.error(f"Error generando bloque: {result2['error']}")
            else:
                herramientas_e2 = result2.get("herramientas_filtradas", [])
                with st.spinner("Enriqueciendo nuevo bloque..."):
                    result3 = call_gemini(model, ETAPA3_PROMPT.format(
                        herramientas_etapa2=json.dumps(herramientas_e2, ensure_ascii=False),
                        tema=tema,
                    ))

                if "error" in result3:
                    st.error(f"Error enriqueciendo bloque: {result3['error']}")
                else:
                    lista_bloque = result3.get("lista_definitiva", [])
                    for i, item in enumerate(lista_bloque):
                        item["numero"] = num_inicio + i
                    st.session_state.all_results.extend(lista_bloque)
                    st.session_state.bloque_num += 1
                    st.rerun()

        # ── Excel download (always accumulated) ──────────────────────────
        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        st.markdown("### 📥 Exportar Resultados")

        excel_bytes = build_excel(all_results, tema)
        col_dl1, col_dl2 = st.columns([1, 3])
        with col_dl1:
            st.download_button(
                label="⬇️ Descargar Excel completo",
                data=excel_bytes,
                file_name=f"buscador_licencias_{tema.lower().replace(' ','_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_dl2:
            st.markdown(
                f'<div class="status-box">Excel con <strong>todos los bloques</strong> '
                f'({len(all_results)} herramientas) en 2 hojas: '
                f'<em>Resultados</em> y <em>Resumen</em>.</div>',
                unsafe_allow_html=True)

        with st.expander("🔧 Ver JSON completo"):
            st.json(all_results)

    elif not run and not st.session_state.e1_done:
        st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
        render_stage_indicator(0)
        st.markdown("""
        <div class="status-box">
        👆 Ingresa un tema y presiona <strong>Buscar Herramientas</strong>.
        Después de la Etapa 1 podrás revisar y corregir antes de continuar.
        Luego podrás generar bloques adicionales sin perder el historial.
        </div>
        """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
